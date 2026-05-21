import os
import sys
import json
from typing import Optional, Union, List, Any
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from sgpi_parser.core.models import ResolucionRectoral, Cronograma, ResultadosConcurso

# Colores del manual de identidad visual de SGPI (Slate-Blue & Teal)
COLOR_TITLE_BG = "1E293B"       # Slate Blue Oscuro (#1E293B)
COLOR_HEADER_BG = "334155"      # Slate Blue Mediano (#334155)
COLOR_ZEBRA_BG = "F8FAFC"       # Gris Azulado Ultra Claro (#F8FAFC)
COLOR_BORDER = "E2E8F0"         # Gris Suave (#E2E8F0)
COLOR_WHITE = "FFFFFF"

def export_data(model: BaseModel, output_path: Optional[str] = None, format_type: str = "json", quiet: bool = False):
    """
    Controlador de exportación de datos. Maneja salidas JSON y Excel.
    """
    format_type = format_type.lower()
    
    if format_type == "json":
        # Generar JSON
        json_data = model.model_dump_json(indent=2)
        
        if output_path:
            out_file = Path(output_path)
            out_file.parent.mkdir(parents=True, exist_ok=True)
            with open(out_file, "w", encoding="utf-8") as f:
                f.write(json_data)
            if not quiet:
                print(f"Éxito: Archivo JSON guardado en: {out_file}")
        else:
            # Si no hay ruta, enviar a stdout
            print(json_data)
            
    elif format_type == "excel":
        if not output_path:
            # Nombre por defecto si no se especifica salida
            category = model.tipo_documento
            year = model.metadata.anio_academico
            output_path = f"{category}_{year}.xlsx"
            
        _export_to_excel(model, output_path, quiet)
    else:
        raise ValueError(f"Formato '{format_type}' no soportado. Elija 'json' o 'excel'.")

def _export_to_excel(model: BaseModel, path: str, quiet: bool):
    """Genera un archivo Excel premium utilizando openpyxl."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos Extraídos"
    
    # Habilitar líneas de cuadrícula visibles
    ws.views.sheetView[0].showGridLines = True
    
    # 1. Definir estilos
    font_main_title = Font(name="Segoe UI", size=15, bold=True, color=COLOR_WHITE)
    font_section_header = Font(name="Segoe UI", size=11, bold=True, color=COLOR_WHITE)
    font_data = Font(name="Segoe UI", size=10)
    font_data_bold = Font(name="Segoe UI", size=10, bold=True)
    
    fill_main_title = PatternFill(start_color=COLOR_TITLE_BG, end_color=COLOR_TITLE_BG, fill_type="solid")
    fill_header = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
    fill_zebra = PatternFill(start_color=COLOR_ZEBRA_BG, end_color=COLOR_ZEBRA_BG, fill_type="solid")
    fill_white = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color=COLOR_BORDER),
        right=Side(style='thin', color=COLOR_BORDER),
        top=Side(style='thin', color=COLOR_BORDER),
        bottom=Side(style='thin', color=COLOR_BORDER)
    )
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_wrap_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 2. Rellenar datos específicos según el modelo
    tipo = model.tipo_documento
    
    if tipo == "cronograma":
        title_text = f"CRONOGRAMA DE ACTIVIDADES: {model.metadata.programa_nombre} ({model.metadata.anio_academico})"
        headers = ["Actividad", "Detalle de Fecha", "Fecha de Inicio", "Fecha de Fin"]
        
        # Escribir Título Principal
        ws.merge_cells("A1:D1")
        ws["A1"] = title_text
        ws["A1"].font = font_main_title
        ws["A1"].fill = fill_main_title
        ws["A1"].alignment = align_center
        ws.row_dimensions[1].height = 40
        
        # Escribir Cabeceras
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = header
            cell.font = font_section_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border
        ws.row_dimensions[2].height = 25
        
        # Escribir Filas de Datos
        row_idx = 3
        for act in model.actividades:
            row_fill = fill_zebra if row_idx % 2 == 1 else fill_white
            
            # Celdas
            c1 = ws.cell(row=row_idx, column=1, value=act.actividad)
            c2 = ws.cell(row=row_idx, column=2, value=act.fecha_detalle)
            c3 = ws.cell(row=row_idx, column=3, value=act.fecha_inicio or "No calculada")
            c4 = ws.cell(row=row_idx, column=4, value=act.fecha_fin or "No calculada")
            
            for c in [c1, c2, c3, c4]:
                c.font = font_data
                c.fill = row_fill
                c.border = thin_border
            
            c1.alignment = align_wrap_left
            c2.alignment = align_left
            c3.alignment = align_center
            c4.alignment = align_center
            
            ws.row_dimensions[row_idx].height = 22
            row_idx += 1
            
    elif tipo == "resultados":
        title_text = f"RESULTADOS DE EVALUACIÓN: {model.metadata.programa_nombre} ({model.metadata.anio_academico})"
        headers = ["Puesto", "Código", "Título del Proyecto", "Responsable", "Facultad", "Nombre de GI", "Puntaje"]
        
        # Título Principal
        ws.merge_cells("A1:G1")
        ws["A1"] = title_text
        ws["A1"].font = font_main_title
        ws["A1"].fill = fill_main_title
        ws["A1"].alignment = align_center
        ws.row_dimensions[1].height = 40
        
        # Cabeceras
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = header
            cell.font = font_section_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border
        ws.row_dimensions[2].height = 25
        
        # Filas de Datos
        row_idx = 3
        for proj in model.proyectos_aprobados:
            row_fill = fill_zebra if row_idx % 2 == 1 else fill_white
            
            c1 = ws.cell(row=row_idx, column=1, value=proj.orden_merito)
            c2 = ws.cell(row=row_idx, column=2, value=proj.codigo_proyecto or "N/A")
            c3 = ws.cell(row=row_idx, column=3, value=proj.titulo)
            c4 = ws.cell(row=row_idx, column=4, value=proj.responsable or "N/A")
            c5 = ws.cell(row=row_idx, column=5, value=proj.facultad or "N/A")
            c6 = ws.cell(row=row_idx, column=6, value=proj.nombre_gi or "N/A")
            c7 = ws.cell(row=row_idx, column=7, value=proj.puntaje)
            
            for c in [c1, c2, c3, c4, c5, c6, c7]:
                c.font = font_data
                c.fill = row_fill
                c.border = thin_border
                
            c1.alignment = align_center
            c2.alignment = align_center
            c3.alignment = align_wrap_left
            c4.alignment = align_left
            c5.alignment = align_left
            c6.alignment = align_left
            
            c7.alignment = align_right
            if proj.puntaje is not None:
                c7.number_format = "0.00"
                
            ws.row_dimensions[row_idx].height = 24
            row_idx += 1
            
    elif tipo == "resolucion_rectoral":
        title_text = f"INTEGRANTES Y PROYECTOS APROBADOS POR {model.metadata.numero_resolucion} ({model.metadata.anio_academico})"
        headers = [
            "Código Proyecto", "Título del Proyecto", "Presupuesto", "Nombre GI Global",
            "Rol Integrante", "Código Miembro", "Apellidos y Nombres", "Tipo Miembro", 
            "Facultad Miembro", "GI Miembro", "Condición GI"
        ]
        
        # Título Principal
        ws.merge_cells("A1:K1")
        ws["A1"] = title_text
        ws["A1"].font = font_main_title
        ws["A1"].fill = fill_main_title
        ws["A1"].alignment = align_center
        ws.row_dimensions[1].height = 40
        
        # Cabeceras
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = header
            cell.font = font_section_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border
        ws.row_dimensions[2].height = 25
        
        # Filas de Datos (Aplanamiento de proyectos e integrantes)
        row_idx = 3
        project_count = 0
        for proj in model.proyectos:
            project_fill = fill_zebra if project_count % 2 == 1 else fill_white
            project_count += 1
            
            # Escribir integrantes
            if not proj.integrantes:
                # Si un proyecto no tiene integrantes (caso raro), escribir una fila vacía para el proyecto
                c_code = ws.cell(row=row_idx, column=1, value=proj.codigo_proyecto)
                c_title = ws.cell(row=row_idx, column=2, value=proj.titulo)
                c_budget = ws.cell(row=row_idx, column=3, value=proj.presupuesto)
                c_gi = ws.cell(row=row_idx, column=4, value=proj.nombre_gi or "N/A")
                
                for c in [c_code, c_title, c_budget, c_gi]:
                    c.font = font_data
                    c.fill = project_fill
                    c.border = thin_border
                c_code.alignment = align_center
                c_title.alignment = align_wrap_left
                c_gi.alignment = align_center
                c_budget.alignment = align_right
                if proj.presupuesto is not None:
                    c_budget.number_format = "$#,##0.00"
                    
                # Rellenar celdas de integrante vacías
                for col in range(5, 12):
                    cell = ws.cell(row=row_idx, column=col, value="-")
                    cell.font = font_data
                    cell.fill = project_fill
                    cell.border = thin_border
                    cell.alignment = align_center
                ws.row_dimensions[row_idx].height = 22
                row_idx += 1
            else:
                # Escribir integrantes del proyecto
                for m_idx, member in enumerate(proj.integrantes):
                    # Columnas de Proyecto
                    # Para el primer integrante escribimos los detalles del proyecto
                    # Para los siguientes podemos dejar en blanco o escribir en gris tenue para no saturar
                    if m_idx == 0:
                        c_code = ws.cell(row=row_idx, column=1, value=proj.codigo_proyecto)
                        c_title = ws.cell(row=row_idx, column=2, value=proj.titulo)
                        c_budget = ws.cell(row=row_idx, column=3, value=proj.presupuesto)
                        c_gi = ws.cell(row=row_idx, column=4, value=proj.nombre_gi or "N/A")
                        
                        for c in [c_code, c_title, c_budget, c_gi]:
                            c.font = font_data_bold
                            c.fill = project_fill
                            c.border = thin_border
                            
                        c_code.alignment = align_center
                        c_title.alignment = align_wrap_left
                        c_gi.alignment = align_center
                        c_budget.alignment = align_right
                        if proj.presupuesto is not None:
                            c_budget.number_format = "$#,##0.00"
                    else:
                        # Para filas siguientes, dejamos en blanco para resaltar la estructura del anexo
                        for col in range(1, 5):
                            cell = ws.cell(row=row_idx, column=col)
                            cell.fill = project_fill
                            cell.border = thin_border

                    # Columnas de Integrante
                    c_rol = ws.cell(row=row_idx, column=5, value=member.rol_proyecto)
                    c_mcode = ws.cell(row=row_idx, column=6, value=member.codigo_miembro or "N/A")
                    c_name = ws.cell(row=row_idx, column=7, value=member.nombre_completo)
                    c_type = ws.cell(row=row_idx, column=8, value=member.tipo_miembro or "N/A")
                    c_fac = ws.cell(row=row_idx, column=9, value=member.facultad or "N/A")
                    c_mgi = ws.cell(row=row_idx, column=10, value=member.gi_codigo or "N/A")
                    c_cond = ws.cell(row=row_idx, column=11, value=member.gi_condicion or "N/A")
                    
                    for c in [c_rol, c_mcode, c_name, c_type, c_fac, c_mgi, c_cond]:
                        c.font = font_data
                        c.fill = project_fill
                        c.border = thin_border
                        
                    c_rol.alignment = align_left
                    c_mcode.alignment = align_center
                    c_name.alignment = align_left
                    c_type.alignment = align_left
                    c_fac.alignment = align_wrap_left
                    c_mgi.alignment = align_center
                    c_cond.alignment = align_center
                    
                    ws.row_dimensions[row_idx].height = 22
                    row_idx += 1

    # 3. Auto-ajustar el ancho de las columnas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Ignorar la primera fila (título principal) en el cálculo de ancho
        for cell in col[1:]:
            val = str(cell.value or '')
            if cell.alignment.wrap_text:
                # Si tiene wrap_text (ej. título largo del proyecto), darle un tamaño fijo y no inflarlo
                max_len = max(max_len, 35)
            else:
                max_len = max(max_len, len(val))
                
        # Agregar margen
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    # 4. Guardar archivo
    output_file = Path(path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    
    if not quiet:
        print(f"Éxito: Reporte de Excel Premium guardado en: {output_file.resolve()}")

from pathlib import Path
